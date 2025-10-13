from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Productos, Proveedor, SalidaProducto, HistorialMovimiento
from django.forms import formset_factory
from django.contrib import messages
from .forms import ProductoForm, MultipleProductosForm, ProveedorForm, SalidaProductoForm, ImportarExcelForm
from decimal import Decimal
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

def inicio(request):
    # Estadísticas de productos
    total_productos = len(Productos.objects.all())
    total_proveedores = len(Proveedor.objects.all())
    productos_bajo_stock = len(Productos.objects.filter(stock__lt=5))
    
    # Productos recientes
    productos_recientes = Productos.objects.order_by('-fecha_ingreso')[:5]
    
    context = {
        'total_productos': total_productos,
        'total_proveedores': total_proveedores,
        'productos_bajo_stock': productos_bajo_stock,
        'productos_recientes': productos_recientes,
    }
    return render(request, 'mi_proyecto/inicio.html', context)

def lista_productos(request):
    categoria_seleccionada = request.GET.get('categoria', 'todos')
    busqueda = request.GET.get('busqueda', '')
    page = request.GET.get('page', 1)
    items_per_page = 10
    
    # Base query con ordenamiento #######################
    productos = Productos.objects.all().order_by('nombre')
    
    # Filtro por categoría #############################
    if categoria_seleccionada != 'todos':
        productos = productos.filter(categoria=categoria_seleccionada)
    
    # Búsqueda por nombre o código ###################
    if busqueda:
        productos = productos.filter(nombre__icontains=busqueda)
    
    # Paginador #############################
    paginator = Paginator(productos, items_per_page)
    
    try:
        productos_paginados = paginator.page(page)
    except PageNotAnInteger:
        productos_paginados = paginator.page(1)
    except EmptyPage:
        productos_paginados = paginator.page(paginator.num_pages)
    
    context = {
        'productos': productos_paginados,
        'categorias': Productos.CATEGORIAS,
        'categoria_actual': categoria_seleccionada,
        'busqueda': busqueda,
    }
    return render(request, 'mi_proyecto/lista_productos.html', context)

def crear_producto(request):
    form = ProductoForm()
    form_multiple = MultipleProductosForm()
    modo_multiple = False
    formset = None

    if request.method == 'POST':
        if 'crear_multiple' in request.POST:
            form_multiple = MultipleProductosForm(request.POST)
            if form_multiple.is_valid():
                cantidad = form_multiple.cleaned_data['cantidad']
                ProductoFormSet = formset_factory(ProductoForm, extra=cantidad)
                formset = ProductoFormSet()
                modo_multiple = True

        elif 'guardar_multiple' in request.POST:
            ProductoFormSet = formset_factory(ProductoForm)
            formset = ProductoFormSet(request.POST)
            if formset.is_valid():
                creados = 0
                for f in formset:
                    if f.has_changed():
                        producto = f.save(commit=False)
                        ######################### Aumento de 30% el precio #########################
                        producto.precio = producto.precio *1.3 
                        producto.save()
                        # Registrar en historial
                        HistorialMovimiento.objects.create(
                            producto=producto,
                            nombre_producto=producto.nombre,
                            serial_producto=producto.codigo,
                            usuario=request.user if request.user.is_authenticated else None,
                            tipo_movimiento='CREACION',
                            detalles='Creación múltiple de productos'
                        )
                        creados += 1
                messages.success(request, f'Se crearon {creados} producto(s).')
                return redirect('lista_productos')
            modo_multiple = True

        else:
            form = ProductoForm(request.POST)
            if form.is_valid():
                producto = form.save(commit=False)
                ######################### Aumento de 30% el precio ################
                producto.precio = producto.precio * Decimal ('1.3')
                producto.precio = producto.precio.quantize(Decimal('0.01'))  
                producto.save()
                # Registrar en historial
                HistorialMovimiento.objects.create(
                    producto=producto,
                    nombre_producto=producto.nombre,
                    serial_producto=producto.codigo,
                    usuario=request.user if request.user.is_authenticated else None,
                    tipo_movimiento='CREACION',
                    detalles='Creación de producto'
                )
                messages.success(request, 'Producto creado exitosamente.')
                return redirect('lista_productos')

    return render(request, 'mi_proyecto/crear_producto.html', {
        'form': form,
        'form_multiple': form_multiple,
        'modo_multiple': modo_multiple,
        'formset': formset
    })

def editar_producto(request, id):
    producto = get_object_or_404(Productos, id=id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            producto = form.save(commit=False)
            ######################### Aumento de 30% el precio #########################
            producto.precio = producto.precio * Decimal('1.3')
            producto.precio = producto.precio.quantize(Decimal('0.01'))
            producto.save()
            # Registrar en historial
            HistorialMovimiento.objects.create(
                producto=producto,
                nombre_producto=producto.nombre,
                serial_producto=producto.codigo,
                usuario=request.user if request.user.is_authenticated else None,
                tipo_movimiento='EDICION',
                detalles='Edición de producto'
            )
            messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente.')
            return redirect('lista_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'mi_proyecto/editar_producto.html', {'form': form, 'producto': producto})

def eliminar_producto(request, id):
    producto = get_object_or_404(Productos, id=id)
    if request.method == 'POST':
        # Guardar datos antes de eliminar
        nombre_producto = producto.nombre
        serial_producto = producto.codigo
        
        # Registrar en historial
        HistorialMovimiento.objects.create(
            producto=None,
            nombre_producto=nombre_producto,
            serial_producto=serial_producto,
            usuario=request.user if request.user.is_authenticated else None,
            tipo_movimiento='ELIMINACION',
            detalles=f'Eliminación de producto: {nombre_producto} (Código: {serial_producto})'
        )
        
        producto.delete()
        return redirect('lista_productos')
    return render(request, 'mi_proyecto/eliminar_producto.html', {'producto': producto}) 


# Proveedor #####################################################

def lista_proveedores(request):
    page = request.GET.get('page', 1)
    items_per_page = 10

    proveedores_qs = Proveedor.objects.all().order_by('nombre')
    paginator = Paginator(proveedores_qs, items_per_page)
    try:
        proveedores = paginator.page(page)
    except PageNotAnInteger:
        proveedores = paginator.page(1)
    except EmptyPage:
        proveedores = paginator.page(paginator.num_pages)

    return render(request, 'mi_proyecto/proveedor/lista_proveedores.html', {
        'proveedores': proveedores
    })

def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm()
    return render(request, 'mi_proyecto/proveedor/crear_proveedor.html', {'form': form})


def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'mi_proyecto/proveedor/editar_proveedor.html', {
        'form': form,
        'proveedor': proveedor
    })

def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    if request.method == 'POST':
        proveedor.delete()
        return redirect('lista_proveedores')
    return render(request, 'mi_proyecto/proveedor/eliminar_proveedor.html', {'proveedor': proveedor})


#Historial de movimientos #####################################################

def historial_movimientos(request):
    categoria_seleccionada = request.GET.get('categoria', 'todos')
    page = request.GET.get('page', 1)
    items_per_page = 10

    movimientos_qs = HistorialMovimiento.objects.select_related('producto', 'usuario').order_by('-fecha_movimiento')
    if categoria_seleccionada != 'todos':
        movimientos_qs = movimientos_qs.filter(producto__categoria=categoria_seleccionada)

    paginator = Paginator(movimientos_qs, items_per_page)
    try:
        movimientos = paginator.page(page)
    except PageNotAnInteger:
        movimientos = paginator.page(1)
    except EmptyPage:
        movimientos = paginator.page(paginator.num_pages)

    return render(request, 'mi_proyecto/historial_movimientos.html', {
        'movimientos': movimientos,
        'categoria_actual': categoria_seleccionada,
        'categorias': Productos.CATEGORIAS,
    })


def lista_salidas(request):
    page = request.GET.get('page', 1)
    items_per_page = 15

    salidas_qs = SalidaProducto.objects.select_related('producto', 'usuario').order_by('-fecha_salida')
    paginator = Paginator(salidas_qs, items_per_page)
    
    try:
        salidas = paginator.page(page)
    except PageNotAnInteger:
        salidas = paginator.page(1)
    except EmptyPage:
        salidas = paginator.page(paginator.num_pages)

    return render(request, 'mi_proyecto/salidas/lista_salidas.html', {
        'salidas': salidas
    })

def registrar_salida(request):
    if request.method == 'POST':
        form = SalidaProductoForm(request.POST)
        if form.is_valid():
            try:
                salida = form.save(commit=False)
                salida.usuario = request.user if request.user.is_authenticated else None
                salida.save()
                messages.success(request, f'Salida registrada: {salida.cantidad} unidades de {salida.producto.nombre}')
                return redirect('lista_salidas')
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = SalidaProductoForm()
    
    return render(request, 'mi_proyecto/salidas/registrar_salidas.html', {'form': form})

def importar_excel(request):
    if request.method == 'POST':
        form = ImportarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            
            try:
                # Cargar el archivo Excel ############################
                workbook = load_workbook(archivo_excel)
                sheet = workbook.active
                
                productos_importados = 0
                productos_actualizados = 0
                errores = []
                
                # Leer las filas del Excel (Las primeras filas son encabezados)
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Extraer datos de cada fila
                        codigo = str(row[0]).strip() if row[0] else None
                        nombre = str(row[1]).strip() if row[1] else None
                        descripcion = str(row[2]).strip() if row[2] else ""
                        precio = float(row[3]) if row[3] else 0
                        stock = int(row[4]) if row[4] else 0
                        categoria = str(row[5]).strip() if row[5] else "COMPUTADORAS"
                        proveedor_nombre = str(row[6]).strip() if row[6] else None
                        
                        # Validar datos requeridos
                        if not nombre or not codigo:
                            errores.append(f"Fila {row_num}: Nombre y código son requeridos")
                            continue
                        
                        # Validar categoría
                        categorias_validas = [cat[0] for cat in Productos.CATEGORIAS]
                        categoria_upper = categoria.upper()
                        if categoria_upper not in categorias_validas:
                            categoria = "COMPUTADORAS"  # Valor por defecto
                        else:
                            categoria = categoria_upper  # Usar la versión en mayúsculas
                        
                        # Buscar o crear proveedor
                        proveedor = None
                        if proveedor_nombre:
                            proveedor, created = Proveedor.objects.get_or_create(
                                nombre=proveedor_nombre,
                                defaults={
                                    'direccion': 'Dirección por definir',
                                    'telefono': '0000-0000',
                                    'email': 'email@ejemplo.com'
                                }
                            )
                        
                        # Aplicar aumento del 30% al precio
                        precio_final = Decimal(str(precio)) * Decimal('1.3')
                        precio_final = precio_final.quantize(Decimal('0.01'))
                        
                        # Verificar si el producto ya existe
                        producto_existente = Productos.objects.filter(codigo=codigo).first()
                        
                        if producto_existente:
                            # Actualizar producto existente
                            producto_existente.nombre = nombre
                            producto_existente.descripcion = descripcion
                            producto_existente.precio = precio_final
                             # Sumar al stock existente 
                            producto_existente.stock += stock  
                            producto_existente.categoria = categoria
                            producto_existente.proveedor = proveedor
                            producto_existente.save()
                            
                            # Registrar en historial
                            HistorialMovimiento.objects.create(
                                producto=producto_existente,
                                nombre_producto=producto_existente.nombre,
                                serial_producto=producto_existente.codigo,
                                usuario=request.user if request.user.is_authenticated else None,
                                tipo_movimiento='EDICION',
                                detalles=f'Actualización desde Excel - Stock agregado: {stock}'
                            )
                            productos_actualizados += 1
                        else:
                            # Crear nuevo producto
                            nuevo_producto = Productos.objects.create(
                                nombre=nombre,
                                codigo=codigo,
                                descripcion=descripcion,
                                precio=precio_final,
                                stock=stock,
                                categoria=categoria,
                                proveedor=proveedor
                            )
                            
                            # Registrar en historial
                            HistorialMovimiento.objects.create(
                                producto=nuevo_producto,
                                nombre_producto=nuevo_producto.nombre,
                                serial_producto=nuevo_producto.codigo,
                                usuario=request.user if request.user.is_authenticated else None,
                                tipo_movimiento='CREACION',
                                detalles='Creación desde Excel'
                            )
                            productos_importados += 1
                            
                    except Exception as e:
                        errores.append(f"Fila {row_num}: Error procesando datos - {str(e)}")
                        continue
                
                # Mostrar resultados
                if productos_importados > 0 or productos_actualizados > 0:
                    mensaje = f"Importación completada: {productos_importados} productos nuevos, {productos_actualizados} productos actualizados."
                    if errores:
                        mensaje += f" Errores: {len(errores)}"
                    messages.success(request, mensaje)
                else:
                    messages.warning(request, "No se importaron productos. Verifique el formato del archivo.")
                
                if errores:
                    for error in errores[:10]:  # Mostrar solo los primeros 10 errores
                        messages.error(request, error)
                
                return redirect('lista_productos')
                
            except Exception as e:
                messages.error(request, f"Error al procesar el archivo Excel: {str(e)}")
    else:
        form = ImportarExcelForm()
    
    return render(request, 'mi_proyecto/importar_excel.html', {'form': form})


def generar_pdf_salida(request, id):
    salida = get_object_or_404(SalidaProducto, id=id)

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    p.setTitle(f"Comprobante de Salida - {salida.producto.nombre}")

    # Encabezado
    p.setFont("Helvetica-Bold", 16)
    p.drawString(72, 750, "Comprobante de Salida de Producto")
    p.setFont("Helvetica", 12)
    p.drawString(72, 730, f"Fecha: {salida.fecha_salida.strftime('%d/%m/%Y %H:%M')}")

    # Información del producto
    p.setFont("Helvetica-Bold", 12)
    p.drawString(72, 700, "Detalles del Producto:")
    p.setFont("Helvetica", 12)
    p.drawString(72, 680, f"Producto: {salida.producto.nombre}")
    p.drawString(72, 660, f"Código: {salida.producto.codigo}")
    p.drawString(72, 640, f"Cantidad: {salida.cantidad}")
    p.drawString(72, 620, f"Motivo: {salida.get_motivo_display()}")

    # Descripción multilínea (si existe)
    if salida.descripcion:
        p.setFont("Helvetica-Bold", 12)
        p.drawString(72, 600, "Descripción:")
        p.setFont("Helvetica", 12)
        y = 580
        for line in salida.descripcion.split('\n'):
            p.drawString(72, y, line)
            y -= 16
            if y < 120:  # salto de página si se acaba el espacio
                p.showPage()
                y = 750
                p.setFont("Helvetica", 12)

    # Usuario
    p.setFont("Helvetica-Bold", 12)
    p.drawString(72, 480, "Registrado por:")
    p.setFont("Helvetica", 12)
    p.drawString(72, 460, f"Usuario: {salida.usuario.username if salida.usuario else 'Sistema'}")

    # Firma
    p.setFont("Helvetica", 10)
    p.drawString(72, 120, "Firma del Responsable: ________________________")

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="salida_{salida.id}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response